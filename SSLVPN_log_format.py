#coding=utf8
import json
import openpyxl
import time
import re
from openpyxl.styles import Font
from nametable import nametable

font = Font(bold=True)

def name_table(nameqp):
    global nametable

    name = {"name":"","department":""}

    if nameqp in nametable.keys():
        name["name"] = nametable[nameqp]["name"]
        name["department"] = nametable[nameqp]["department"]
    else:
        name["name"] = nameqp

    return name

def write_rows(ws,datalist):
    row = 2
    for eachrow in datalist:
        for col in range(1,7):
            ws.cell(column=col,row=row,value=eachrow[col-1])
        row += 1
        print("["+ str(eachrow[1]) +"] "+ str(eachrow[2]) + ' ' + str(eachrow[3]) + ',IP: ' + str(eachrow[4]))
            

if __name__ == '__main__':

    with open("logdata.txt","r",encoding="utf8") as f:
        logdata = f.read()

    logdata_json = json.loads(logdata)
    datas = logdata_json["result"]["data"]

    timeinit = time.localtime(time.time())
    date = str(timeinit.tm_mon) + '月' + str(timeinit.tm_mday) + '日'


    wb = openpyxl.load_workbook(filename="VPN登录日志.xlsx")
    ws = wb.create_sheet(date)

    row0 = ['登录日期','登录时间','姓名','动作','登录IP','部门']
    for col in range(1,7):
        ws.cell(column=col,row=1,value=row0[col-1])

    
    datalist = []
    for item in datas:
        if ("User" in item["detailinfo"]) and ("MOD_TWF" in item["detailinfo"]):
            if ("login" in item["detailinfo"]):
                logtime = item["logdate"]
                detailinfo = item["detailinfo"]
                nameqp = re.search(r'User (\S+) login',detailinfo).group(1)
                dic = name_table(nameqp)
                department = dic["department"]
                name = dic["name"]
                movement = "登录"
                IP = ""
                VPN = "深信服VPN"
                datalist.append([date,logtime,name,movement,IP,department])
                
            elif ("logout" in item["detailinfo"]):
                logtime = item["logdate"]
                detailinfo = item["detailinfo"]
                nameqp = re.search(r'User (\S+) logout',detailinfo).group(1)
                dic = name_table(nameqp)
                department = dic["department"]
                name = dic["name"]
                movement = "退出"
                IP = re.search(r'ip address: (\d+\.\d+\.\d+\.\d+),',detailinfo).group(1)
                VPN = "深信服VPN"
                datalist.append([date,logtime,name,movement,IP,department])

    datalist.reverse()
    write_rows(ws,datalist)


    wb.save(filename="VPN登录日志.xlsx")